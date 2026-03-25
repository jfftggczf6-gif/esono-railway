"""
cell_filler.py — Moteur générique de remplissage Excel.
Aucune logique métier. Reçoit {"SheetName!CellRef": value}, écrit dans le template.
"""
import openpyxl
import io
import base64
import os
from fastapi import Request, HTTPException
from fastapi.responses import Response
from typing import Optional
from fastapi import Header


def register_cell_filler(app):
    @app.post("/fill-cells")
    async def fill_cells(request: Request, authorization: Optional[str] = Header(None)):
        PK = os.getenv("PARSER_API_KEY", "esono-parser-dev-key")
        if authorization != f"Bearer {PK}":
            raise HTTPException(status_code=401, detail="Invalid API key")

        body = await request.json()
        cell_values = body.get("cell_values", {})
        t64 = body.get("template_base64")

        if not t64:
            raise HTTPException(status_code=400, detail="template_base64 required")
        if not cell_values:
            raise HTTPException(status_code=400, detail="cell_values required")

        tb = base64.b64decode(t64)
        wb = openpyxl.load_workbook(io.BytesIO(tb), keep_vba=True, data_only=False)

        filled = 0
        skipped_formula = 0
        skipped_error = 0

        for ref, value in cell_values.items():
            try:
                if "!" in ref:
                    sheet_name, cell_ref = ref.split("!", 1)
                else:
                    sheet_name = "InputsData"
                    cell_ref = ref

                if sheet_name not in wb.sheetnames:
                    skipped_error += 1
                    continue

                ws = wb[sheet_name]
                cell = ws[cell_ref]

                # Skip formulas
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    skipped_formula += 1
                    continue

                # Type coercion
                if isinstance(value, str):
                    try:
                        value = int(value)
                    except ValueError:
                        try:
                            value = float(value)
                        except ValueError:
                            pass

                cell.value = value
                filled += 1
            except Exception as e:
                print(f"[fill-cells] Error {ref}={value}: {e}")
                skipped_error += 1

        # Force recalc on open
        wb.calculation = openpyxl.workbook.properties.CalcProperties(fullCalcOnLoad=True)

        print(f"[fill-cells] Done: {filled} filled, {skipped_formula} formula-skipped, {skipped_error} errors")

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        return Response(
            content=out.read(),
            media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
            headers={"Content-Disposition": "attachment; filename=PlanFinancierOVO.xlsm"}
        )
