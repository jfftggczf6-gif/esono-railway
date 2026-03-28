"""
ovo_filler.py v2 — Remplissage direct du template OVO depuis le JSON plan_financier.

Basé sur l'analyse du fichier GOTCHE de référence (bien rempli).
Principe: force-write (fw) partout sauf les lignes de TOTAUX et les feuilles Pivot/Chart.

Source: plan_financier JSON (produits, staff, opex, capex, loans, projections, etc.)
Target: 251022-PlanFinancierOVO-Template5Ans-v0210-EMPTY.xlsm
"""

import openpyxl
from datetime import datetime
import math

# ═══════════════════════════════════════════════════════════════
# CONSTANTES
# ═══════════════════════════════════════════════════════════════

YEAR_LABEL = {
    "YEAR-2": "year_minus_2", "YEAR-1": "year_minus_1",
    "CURRENT YEAR": "current_year", "CURRENT YEAR H1": "h1", "CURRENT YEAR H2": "h2",
    "YEAR2": "year2", "YEAR3": "year3", "YEAR4": "year4", "YEAR5": "year5", "YEAR6": "year6",
}

FIN_COL = {
    "year_minus_2": "O", "year_minus_1": "P", "h1": "Q", "h2": "R",
    "current_year": "S", "year2": "T", "year3": "U", "year4": "V",
    "year5": "W", "year6": "X",
}

IDX_YEAR = ["year_minus_2", "year_minus_1", "h1", "h2", "current_year",
            "year2", "year3", "year4", "year5", "year6"]

REV_OFFSET = {
    "year_minus_2": 0, "year_minus_1": 1, "current_year": 2,
    "year2": 3, "year3": 4, "year4": 5, "year5": 6, "year6": 7,
}

PRODUCT_BLOCK = 42   # 8 data rows + 34 calc rows per product
PRODUCT_START = 9    # Row 9 = Product 1 YEAR-2
SERVICE_START = 849  # Row 849 = Service 1 YEAR-2

STAFF_ROWS = {
    0: {"eft": 213, "sal": 214, "allow": 215, "ss": 216},
    1: {"eft": 220, "sal": 221, "allow": 222, "ss": 223},
    2: {"eft": 227, "sal": 228, "allow": 229, "ss": 230},
    3: {"eft": 234, "sal": 235, "allow": 236, "ss": 237},
    4: {"eft": 241, "sal": 242, "allow": 243, "ss": 244},
    5: {"eft": 248, "sal": 249, "allow": 250, "ss": 251},
    6: {"eft": 255, "sal": 256, "allow": 257, "ss": 258},
    7: {"eft": 262, "sal": 263, "allow": 264, "ss": 265},
    8: {"eft": 269, "sal": 270, "allow": 271, "ss": 272},
    9: {"eft": 276, "sal": 277, "allow": 278, "ss": 279},
}

CAPEX_SLOTS = {
    "office":      {"start": 408, "count": 7, "amort": 0.25},
    "production":  {"start": 450, "count": 7, "amort": 0.10},
    "equipment":   {"start": 450, "count": 7, "amort": 0.10},
    "other":       {"start": 462, "count": 7, "amort": 0.10},
    "commercial":  {"start": 462, "count": 7, "amort": 0.10},
    "intangible":  {"start": 486, "count": 5, "amort": 0.20},
    "incorporel":  {"start": 486, "count": 5, "amort": 0.20},
    "logiciel":    {"start": 486, "count": 5, "amort": 0.20},
    "startup":     {"start": 493, "count": 5, "amort": 0.20},
    "demarrage":   {"start": 493, "count": 5, "amort": 0.20},
    "technologie": {"start": 486, "count": 5, "amort": 0.33},
    "logistique":  {"start": 462, "count": 7, "amort": 0.20},
    "vehicule":    {"start": 462, "count": 7, "amort": 0.20},
    "transport":   {"start": 462, "count": 7, "amort": 0.20},
}

OPEX_MARKETING = {"research": 201, "studies": 202, "receptions": 203,
                  "documentation": 204, "advertising": 205}
OPEX_OFFICE = {"rent": 294, "internet": 295, "telecom": 296, "supplies": 297,
               "equipment": 298, "packaging": 299, "fuel": 300, "water": 301,
               "electricity": 302, "cleaning": 303}
OPEX_OTHER = {"health": 311, "directors": 312, "donations": 313, "sponsorship": 314}
OPEX_INSURANCE = {"building": 326, "company": 327}
OPEX_MAINTENANCE = {"movable": 335, "demolition": 336, "other": 337}
OPEX_THIRD = {"legal_startup": 345, "loan_followup": 346, "other_startup": 347,
              "transport": 348, "delivery": 349, "commissions": 350,
              "legal_adviser": 351, "accounting": 352}

SCENARIO_ROWS = {
    "revenue_products": 130, "cogs_products": 131,
    "revenue_services": 132, "cogs_services": 133,
    "marketing": 134, "staff_salaries": 135, "taxes_staff": 136,
    "office": 137, "other_expenses": 138, "travel": 139,
    "insurance": 140, "maintenance": 141, "third_parties": 142,
}

LOAN_YEAR_COLS = {"current_year": "S", "year2": "T", "year3": "U",
                  "year4": "V", "year5": "W", "year6": "X"}


# ═══════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════

def fw(ws, row, col, value):
    """Force write: écrase même les formules."""
    ws[f"{col}{row}"].value = value


def arr_to_dict(arr):
    """Array 10 valeurs → dict {year_key: value}."""
    if isinstance(arr, dict):
        return arr
    if not isinstance(arr, list):
        return {}
    return {IDX_YEAR[i]: (arr[i] or 0) for i in range(min(len(arr), 10))}


def write_fin_row(ws, row, data):
    """Écrit une ligne FinanceData. Force-write sur TOUTES les colonnes O-X.
    Si CY (index 4) = 0 mais H1 et H2 ont des valeurs → CY = H1 + H2."""
    d = arr_to_dict(data) if isinstance(data, (list, dict)) else {}
    # Fix CY=0 quand H1+H2 existent
    h1 = d.get("h1", 0) or 0
    h2 = d.get("h2", 0) or 0
    cy = d.get("current_year", 0) or 0
    if cy == 0 and (h1 + h2) > 0:
        d["current_year"] = h1 + h2
    for yk, col in FIN_COL.items():
        v = d.get(yk)
        if v is not None:
            fw(ws, row, col, v)


def get_par_annee_dict(item):
    """par_annee[] → {year_key: {fields...}}."""
    result = {}
    for pa in item.get("par_annee", []):
        label = pa.get("annee", "")
        yk = YEAR_LABEL.get(label)
        if yk:
            result[yk] = pa
    return result


# ═══════════════════════════════════════════════════════════════
# FILL PRINCIPAL
# ═══════════════════════════════════════════════════════════════

def fill_plan_financier(wb, data):
    inp = wb["InputsData"]
    rev = wb["RevenueData"]
    fin = wb["FinanceData"]
    rdm = wb["ReadMe"]

    fw(rdm, 3, "L", "French")

    # ═══ 1. InputsData — Company ═══
    fw(inp, 5, "J", data.get("company", ""))
    fw(inp, 6, "J", data.get("country", ""))
    fw(inp, 8, "J", data.get("currency", ""))
    fw(inp, 9, "J", data.get("exchange_rate_eur", 655.957))
    fw(inp, 10, "J", datetime.now())
    fw(inp, 12, "J", data.get("vat_rate", 0.18))
    fw(inp, 14, "J", data.get("inflation_rate", 0.03))
    fw(inp, 17, "J", data.get("tax_regime_1", 0.04))
    fw(inp, 18, "J", data.get("tax_regime_2", 0.25))

    # ═══ 2. Years ═══
    years = data.get("years", {})
    ym2 = years.get("year_minus_2", data.get("current_year", 2024) - 2)
    fw(inp, 24, "J", ym2)
    fw(inp, 25, "J", ym2 + 1)
    fw(inp, 26, "J", ym2 + 2)  # CY
    fw(inp, 27, "J", ym2 + 2)  # H1
    fw(inp, 28, "J", ym2 + 2)  # H2
    fw(inp, 29, "J", ym2 + 3)
    fw(inp, 30, "J", ym2 + 4)
    fw(inp, 31, "J", ym2 + 5)
    fw(inp, 32, "J", ym2 + 6)
    fw(inp, 33, "J", ym2 + 7)

    # ═══ 3. Products & Services names ═══
    produits = data.get("produits", [])
    services = data.get("services", [])

    for i in range(20):
        row = 36 + i
        if i < len(produits):
            fw(inp, row, "H", produits[i].get("nom", "-"))
            fw(inp, row, "I", 1)
        else:
            fw(inp, row, "H", "-")
            fw(inp, row, "I", 0)

    for i in range(10):
        row = 58 + i
        if i < len(services):
            fw(inp, row, "H", services[i].get("nom", "-"))
            fw(inp, row, "I", 1)
        else:
            fw(inp, row, "H", "-")
            fw(inp, row, "I", 0)

    # ═══ 4. Ranges & Channels ═══
    ranges = data.get("ranges", [])
    if isinstance(ranges, list):
        fw(inp, 70, "J", ranges[0].get("name", "STANDARD") if len(ranges) > 0 else "STANDARD")
        fw(inp, 71, "J", ranges[1].get("name", "-") if len(ranges) > 1 else "-")
        fw(inp, 72, "J", ranges[2].get("name", "-") if len(ranges) > 2 else "-")
    channels = data.get("channels", [])
    if isinstance(channels, list):
        fw(inp, 75, "J", channels[0].get("name", "B2B") if len(channels) > 0 else "B2B")
        fw(inp, 76, "J", channels[1].get("name", "B2C") if len(channels) > 1 else "B2C")

    # Product × Range/Channel matrix
    for i in range(len(produits[:20])):
        row = 79 + i
        fw(inp, row, "F", 1); fw(inp, row, "G", 0); fw(inp, row, "H", 0)
        fw(inp, row, "I", 0); fw(inp, row, "J", 1)
    for i in range(len(services[:10])):
        row = 101 + i
        fw(inp, row, "F", 1); fw(inp, row, "G", 0); fw(inp, row, "H", 0)
        fw(inp, row, "I", 0); fw(inp, row, "J", 1)

    # ═══ 5. RevenueData — Products & Services ═══
    def fill_revenue_item(base_row, item):
        """Remplit un bloc de 8 lignes. Force-write TOUT comme GOTCHE."""
        pa_dict = get_par_annee_dict(item)
        nom = item.get("nom", "")
        fallback_price = item.get("prix_unitaire", 0) or 0
        fallback_cogs = item.get("cout_unitaire", 0) or 0

        for yk, offset in REV_OFFSET.items():
            row = base_row + offset
            yr = pa_dict.get(yk, {})

            prix_r1 = yr.get("prix_r1", fallback_price) or fallback_price
            prix_r2 = yr.get("prix_r2", 0) or 0
            prix_r3 = yr.get("prix_r3", 0) or 0
            cogs_r1 = yr.get("cogs_r1", fallback_cogs) or fallback_cogs
            cogs_r2 = yr.get("cogs_r2", 0) or 0
            cogs_r3 = yr.get("cogs_r3", 0) or 0
            mix_r1 = yr.get("mix_r1", 1.0)
            mix_r2 = yr.get("mix_r2", 0)
            mix_ch1 = yr.get("mix_ch1", 1.0)

            vol = yr.get("volume", 0) or 0
            q1 = yr.get("volume_q1", 0) or round(vol * 0.22)
            q2 = yr.get("volume_q2", 0) or round(vol * 0.25)
            q3 = yr.get("volume_q3", 0) or round(vol * 0.27)
            q4 = yr.get("volume_q4", 0) or (vol - q1 - q2 - q3)
            vol_total = q1 + q2 + q3 + q4 if (q1 + q2 + q3 + q4) > 0 else vol

            # Col J: nom du produit (comme GOTCHE)
            fw(rev, row, "J", nom)

            # Prix: L=R1, M=R2, N=R3, O=copie de L (force)
            fw(rev, row, "L", prix_r1)
            fw(rev, row, "M", prix_r2)
            fw(rev, row, "N", prix_r3)
            fw(rev, row, "O", prix_r1)

            # Mix volume: P=R1, Q=R2, R=0 (force)
            fw(rev, row, "P", mix_r1)
            fw(rev, row, "Q", mix_r2)
            fw(rev, row, "R", 0)

            # COGS: S=R1, T=R2, U=R3, V=copie de S (force)
            fw(rev, row, "S", cogs_r1)
            fw(rev, row, "T", cogs_r2)
            fw(rev, row, "U", cogs_r3)
            fw(rev, row, "V", cogs_r1)

            # Channel mix: W=ch1, X=0, Y=0 (force)
            fw(rev, row, "W", mix_ch1)
            fw(rev, row, "X", 0)
            fw(rev, row, "Y", 0)
            fw(rev, row, "Z", 0)
            fw(rev, row, "AA", 0)
            fw(rev, row, "AB", 0)

            # Volumes: TOUJOURS force-write Q1-Q4 ET total
            fw(rev, row, "AE", round(q1))
            fw(rev, row, "AF", round(q2))
            fw(rev, row, "AG", round(q3))
            fw(rev, row, "AH", round(q4))
            fw(rev, row, "AI", round(vol_total))

    for i, p in enumerate(produits[:20]):
        base = PRODUCT_START + i * PRODUCT_BLOCK
        print(f"[ovo_filler] Product {i+1} '{p.get('nom', '?')[:30]}' → rows {base}-{base+7}")
        fill_revenue_item(base, p)

    for i, s in enumerate(services[:10]):
        base = SERVICE_START + i * PRODUCT_BLOCK
        print(f"[ovo_filler] Service {i+1} '{s.get('nom', '?')[:30]}' → rows {base}-{base+7}")
        fill_revenue_item(base, s)

    # ═══ 6. Staff ═══
    staff_list = data.get("staff", [])

    for i in range(10):
        row = 113 + i
        if i < len(staff_list):
            s = staff_list[i]
            fw(inp, row, "H", s.get("categorie", ""))
            fw(inp, row, "I", s.get("departement", ""))
            fw(inp, row, "J", s.get("taux_charges_sociales", 0.185))
        else:
            fw(inp, row, "J", 0)

    for i, s in enumerate(staff_list[:10]):
        rows = STAFF_ROWS.get(i)
        if not rows:
            continue
        pa_dict = get_par_annee_dict(s)

        # Force-write sur TOUTES les colonnes O-X y compris S (comme GOTCHE)
        for yk, col in FIN_COL.items():
            yr = pa_dict.get(yk, {})
            if not isinstance(yr, dict):
                continue
            fw(fin, rows["eft"], col, round(yr.get("effectif", 0)))
            fw(fin, rows["sal"], col, yr.get("salaire_mensuel_brut", 0))
            fw(fin, rows["allow"], col, yr.get("primes_annuelles", 0))
            fw(fin, rows["ss"], col, yr.get("charges_sociales_mensuelles", 0))

    # ═══ 7. Loans ═══
    loans = data.get("loans", {})
    ovo_loan = loans.get("ovo", {})
    fam_loan = loans.get("family", {})
    bnk_loan = loans.get("bank", {})

    fw(inp, 125, "I", ovo_loan.get("rate", ovo_loan.get("interest_rate", 0.07)))
    fw(inp, 125, "J", ovo_loan.get("term_years", ovo_loan.get("duration_years", 5)))
    fw(inp, 126, "I", fam_loan.get("rate", fam_loan.get("interest_rate", 0.10)))
    fw(inp, 126, "J", fam_loan.get("term_years", fam_loan.get("duration_years", 3)))
    fw(inp, 127, "I", bnk_loan.get("rate", bnk_loan.get("interest_rate", 0.12)))
    fw(inp, 127, "J", bnk_loan.get("term_years", bnk_loan.get("duration_years", 2)))

    # Loans: disbursement (785-787), grace (793/802/811), repayments, interest
    for loan_data, disb_row, grace_row, rate_row, interest_h1_row in [
        (ovo_loan, 785, 793, 797, 798),
        (fam_loan, 786, 802, 806, 807),
        (bnk_loan, 787, 811, 815, 816),
    ]:
        amount = loan_data.get("amount", 0) or 0
        rate = loan_data.get("rate", loan_data.get("interest_rate", 0)) or 0
        term = max(1, loan_data.get("term_years", loan_data.get("duration_years", 5)) or 5)

        if amount > 0:
            # Disbursement dans T (Y2)
            fw(fin, disb_row, "T", amount)
            # Aussi écrire dans J pour la référence
            fw(fin, disb_row, "J", amount)

            # Durée
            fw(fin, grace_row, "J", term)

            # Grace periods: 0 partout sauf les années après la fin du prêt
            year_keys = list(LOAN_YEAR_COLS.keys())
            for j, yk in enumerate(year_keys):
                col = LOAN_YEAR_COLS[yk]
                fw(fin, grace_row, col, 0 if j < term else 0)

            # Repayments per year (rows grace_row+1, grace_row+2)
            yearly_repay = round(amount / term)
            remaining = amount
            for j, yk in enumerate(year_keys):
                col = LOAN_YEAR_COLS[yk]
                if j < term and remaining > 0:
                    repay = min(yearly_repay, remaining)
                    fw(fin, grace_row + 1, col, repay)   # Capital repayment
                    fw(fin, grace_row + 2, col, repay)   # Total repayment
                    remaining -= repay

            # Interest rate per year
            fw(fin, rate_row, "J", rate)
            fw(fin, rate_row, "S", rate)
            for col in LOAN_YEAR_COLS.values():
                fw(fin, rate_row, col, rate)

            # Interest H1 current year
            fw(fin, interest_h1_row, "S", round(amount * rate / 2))

        else:
            for col in LOAN_YEAR_COLS.values():
                fw(fin, grace_row, col, 0)

    # Existing/new shareholders
    financing = data.get("financing", {})
    all_cols = {"year_minus_2": "O", "year_minus_1": "P", **LOAN_YEAR_COLS}
    ex_sh = financing.get("existing_shareholders", {})
    if isinstance(ex_sh, dict):
        for yk, col in all_cols.items():
            fw(fin, 788, col, ex_sh.get(yk, 0))
    nw_sh = financing.get("new_shareholders", {})
    if isinstance(nw_sh, dict):
        for yk, col in LOAN_YEAR_COLS.items():
            fw(fin, 789, col, nw_sh.get(yk, 0))

    # ═══ 8. Scenarios (multiplicateurs, pas taux de croissance) ═══
    hyp = data.get("hypotheses_ia", {})
    growth_rates = hyp.get("taux_croissance_ca", [0.15])
    avg_growth = sum(growth_rates) / len(growth_rates) if growth_rates else 0.15

    scenarios = data.get("scenarios", {})
    if isinstance(scenarios, dict):
        first_val = next(iter(scenarios.values()), None)
        if isinstance(first_val, dict) and "worst" in first_val:
            for key, row in SCENARIO_ROWS.items():
                s = scenarios.get(key, {})
                if isinstance(s, dict):
                    if s.get("worst") is not None: fw(inp, row, "H", s["worst"])
                    if s.get("typical") is not None: fw(inp, row, "I", s["typical"])
                    if s.get("best") is not None: fw(inp, row, "J", s["best"])
        else:
            _fill_default_scenarios(inp, avg_growth)
    else:
        _fill_default_scenarios(inp, avg_growth)

    # ═══ 9. OPEX (force-write O-X) ═══
    opex = data.get("opex", {})

    mkt = opex.get("marketing", {})
    for key, row in OPEX_MARKETING.items():
        write_fin_row(fin, row, mkt.get(key, {}))
    for i, x in enumerate(mkt.get("extra", [])[:5]):
        write_fin_row(fin, 206 + i, x)

    tx = opex.get("taxes_on_staff", opex.get("taxes_staff", {}))
    write_fin_row(fin, 283, tx.get("salaries_tax", {}))
    write_fin_row(fin, 284, tx.get("apprenticeship", {}))
    write_fin_row(fin, 285, tx.get("training", {}))
    write_fin_row(fin, 286, tx.get("other", {}))
    for i, x in enumerate(tx.get("extra", [])[:5]):
        write_fin_row(fin, 287 + i, x)

    of = opex.get("office", {})
    for key, row in OPEX_OFFICE.items():
        write_fin_row(fin, row, of.get(key, {}))
    for i, x in enumerate(of.get("extra", [])[:5]):
        write_fin_row(fin, 304 + i, x)

    ot = opex.get("other", opex.get("other_expenses", {}))
    for key, row in OPEX_OTHER.items():
        write_fin_row(fin, row, ot.get(key, {}))
    for i, x in enumerate(ot.get("extra", [])[:5]):
        write_fin_row(fin, 315 + i, x)

    tr = opex.get("travel", {})
    write_fin_row(fin, 322, tr.get("nb_travellers", tr.get("headcount", {})))
    write_fin_row(fin, 323, tr.get("avg_cost", tr.get("cost_per_person", {})))

    ins = opex.get("insurance", {})
    for key, row in OPEX_INSURANCE.items():
        write_fin_row(fin, row, ins.get(key, {}))
    for i, x in enumerate(ins.get("extra", [])[:5]):
        write_fin_row(fin, 328 + i, x)

    mn = opex.get("maintenance", {})
    for key, row in OPEX_MAINTENANCE.items():
        write_fin_row(fin, row, mn.get(key, {}))
    for i, x in enumerate(mn.get("extra", [])[:5]):
        write_fin_row(fin, 338 + i, x)

    tp = opex.get("third_parties", {})
    for key, row in OPEX_THIRD.items():
        write_fin_row(fin, row, tp.get(key, {}))
    for i, x in enumerate(tp.get("extra", [])[:5]):
        write_fin_row(fin, 353 + i, x)

    # ═══ 10. Stock Variation ═══
    sv = data.get("stock_variation", {})
    write_fin_row(fin, 188, sv.get("increase", {}))
    write_fin_row(fin, 189, sv.get("decrease", {}))

    # ═══ 11. CAPEX (col H pour labels, comme GOTCHE) ═══
    capex_raw = data.get("capex", [])
    if isinstance(capex_raw, list):
        categorized = {}
        for item in capex_raw:
            if not isinstance(item, dict):
                continue
            cat = (item.get("categorie") or item.get("category") or item.get("type") or "other").lower()
            if cat in ("bureau", "office_equipment"): cat = "office"
            elif cat in ("production_equipment",): cat = "production"
            elif cat in ("incorporel", "logiciel", "software"): cat = "intangible"
            elif cat in ("demarrage",): cat = "startup"
            elif cat not in CAPEX_SLOTS: cat = "other"
            categorized.setdefault(cat, []).append(item)
    elif isinstance(capex_raw, dict):
        categorized = capex_raw
    else:
        categorized = {}

    for cat, items in categorized.items():
        slot = CAPEX_SLOTS.get(cat)
        if not slot or not isinstance(items, list):
            continue
        for j, item in enumerate(items[:slot["count"]]):
            row = slot["start"] + j
            fw(fin, row, "H", item.get("label", item.get("nature", "")))
            fw(fin, row, "K", item.get("acquisition_year", item.get("annee", 0)))
            fw(fin, row, "L", item.get("acquisition_value", item.get("montant", item.get("value", 0))))
            fw(fin, row, "M", item.get("amortisation_rate", item.get("taux_amortissement",
                                       item.get("amort_rate", slot["amort"]))))

    # ═══ 12. Working Capital ═══
    wc = data.get("working_capital", {})
    bfr = data.get("bfr_detail", {})

    for wc_key, bfr_key, fin_row in [
        ("stock_days", "stock_moyen_jours", 693),
        ("receivable_days", "delai_clients_jours", 697),
        ("payable_days", "delai_fournisseurs_jours", 701),
    ]:
        vals = wc.get(wc_key, [])
        if vals:
            write_fin_row(fin, fin_row, vals)
        elif bfr and bfr.get(bfr_key):
            val = bfr[bfr_key]
            for col in FIN_COL.values():
                fw(fin, fin_row, col, val)

    # ═══ 13. P&L Adjustments ═══
    adj = data.get("adjustments", {})

    # Bank charges (row 729) — toutes colonnes
    bc = adj.get("bank_charges", {})
    if isinstance(bc, dict) and bc:
        write_fin_row(fin, 729, bc)
    elif isinstance(bc, (int, float)) and bc > 0:
        for col in FIN_COL.values():
            fw(fin, 729, col, bc)

    # Intérêts: calculer depuis loans si pas dans adjustments
    for loan_data, adj_key, fin_row in [
        (ovo_loan, "interest_ovo", 730),
        (fam_loan, "interest_family", 731),
        (bnk_loan, "interest_bank", 732),
    ]:
        d = adj.get(adj_key, {})
        amount = loan_data.get("amount", 0) or 0
        rate = loan_data.get("rate", loan_data.get("interest_rate", 0)) or 0
        term = max(1, loan_data.get("term_years", loan_data.get("duration_years", 5)) or 5)

        if isinstance(d, dict) and d:
            write_fin_row(fin, fin_row, d)
        elif amount > 0 and rate > 0:
            yearly_repay = round(amount / term)
            remaining = amount
            # Historique
            fw(fin, fin_row, "O", 0)
            fw(fin, fin_row, "P", 0)
            # CY et futures: intérêts décroissants
            for yk, col in LOAN_YEAR_COLS.items():
                interest = round(remaining * rate)
                fw(fin, fin_row, col, interest)
                remaining = max(0, remaining - yearly_repay)

    # Tax regimes
    taxes = adj.get("taxes", {})
    write_fin_row(fin, 737, taxes.get("regime_1", {}))
    write_fin_row(fin, 738, taxes.get("regime_2", {}))

    # ═══ 14. Cash Position ═══
    kpis = data.get("kpis", {})
    projections = data.get("projections", [])

    cash_ym2, cash_ym1 = 0, 0
    for p in projections:
        if p.get("annee") == "YEAR-2":
            cash_ym2 = p.get("cashflow", 0)
        elif p.get("annee") == "YEAR-1":
            cash_ym1 = p.get("cashflow", 0)
    if kpis.get("tresorerie"):
        cash_ym1 = kpis["tresorerie"]

    fw(fin, 773, "O", cash_ym2)
    fw(fin, 773, "P", cash_ym1)

    # ═══ 15. P&L TOTAUX — écraser les GETPIVOTDATA (comme GOTCHE) ═══
    # projections[] contient: ca, cogs, marge_brute, opex_total, ebitda, amortissements,
    #                         charges_financieres, resultat_exploitation, impots, resultat_net, cashflow
    projections = data.get("projections", [])

    # Map année label → year_key pour FIN_COL
    PROJ_YEAR_MAP = {
        "YEAR-2": "year_minus_2", "YEAR-1": "year_minus_1",
        "CURRENT YEAR": "current_year",
        "YEAR2": "year2", "YEAR3": "year3", "YEAR4": "year4",
        "YEAR5": "year5", "YEAR6": "year6",
    }

    # Construire dict {year_key: projection_data}
    proj_by_year = {}
    for p in projections:
        yk = PROJ_YEAR_MAP.get(p.get("annee", ""))
        if yk:
            proj_by_year[yk] = p

    def write_proj_row(fin_row, field, default=0):
        """Écrit une ligne P&L depuis projections."""
        for yk, col in FIN_COL.items():
            proj = proj_by_year.get(yk, {})
            val = proj.get(field, default)
            if val is not None:
                fw(fin, fin_row, col, val)

    # Revenue per product (rows 52+i pour chaque produit, 72=total produits)
    # On calcule le CA par produit depuis produits[].par_annee[].ca
    for i, p in enumerate(produits[:20]):
        pa_dict = get_par_annee_dict(p)
        row_rev = 52 + i   # Revenue product i
        row_cogs = 96 + i  # COGS product i
        row_gp = 118 + i   # Gross profit product i
        for yk, col in FIN_COL.items():
            yr = pa_dict.get(yk, {})
            ca = yr.get("ca", 0) or 0
            cogs = yr.get("cogs_total", 0) or 0
            # Si pas de ca calculé, estimer depuis prix × volume
            if ca == 0:
                vol = yr.get("volume", 0) or 0
                prix = yr.get("prix_r1", p.get("prix_unitaire", 0)) or 0
                ca = vol * prix
            if cogs == 0:
                vol = yr.get("volume", 0) or 0
                cogs_u = yr.get("cogs_r1", p.get("cout_unitaire", 0)) or 0
                cogs = vol * cogs_u
            gp = ca - cogs
            fw(fin, row_rev, col, ca)
            fw(fin, row_cogs, col, cogs)
            fw(fin, row_gp, col, gp)

    # Volume total par produit (row 8+i)
    for i, p in enumerate(produits[:20]):
        pa_dict = get_par_annee_dict(p)
        for yk, col in FIN_COL.items():
            yr = pa_dict.get(yk, {})
            vol = yr.get("volume", 0) or 0
            fw(fin, 8 + i, col, vol)

    # Prix moyen par produit (row 30+i)
    for i, p in enumerate(produits[:20]):
        pa_dict = get_par_annee_dict(p)
        for yk, col in FIN_COL.items():
            yr = pa_dict.get(yk, {})
            prix = yr.get("prix_r1", p.get("prix_unitaire", 0)) or 0
            fw(fin, 30 + i, col, prix)

    # COGS unitaire moyen (row 74+i)
    for i, p in enumerate(produits[:20]):
        pa_dict = get_par_annee_dict(p)
        for yk, col in FIN_COL.items():
            yr = pa_dict.get(yk, {})
            cogs_u = yr.get("cogs_r1", p.get("cout_unitaire", 0)) or 0
            fw(fin, 74 + i, col, cogs_u)

    # Totaux consolidés (comme GOTCHE)
    # Row 72 = Revenue Products TOTAL, 116 = COGS Products TOTAL, 138 = GP Products TOTAL
    # Row 140 = Revenue Products TOTAL (bis), 192 = Revenue TOTAL, 194 = COGS TOTAL, 195 = GP TOTAL
    # Row 710 = Revenue TOTAL (P&L), 712 = COGS, 714 = GP, 716 = OPEX, 718 = EBITDA
    # Row 720 = Amortisations, 722 = EBIT, 735 = Profit before tax, 742 = Taxes, 744 = Net profit
    write_proj_row(72, "ca")       # Revenue Products TOTAL
    write_proj_row(116, "cogs")    # COGS Products TOTAL
    # GP Products = CA - COGS
    for yk, col in FIN_COL.items():
        proj = proj_by_year.get(yk, {})
        ca = proj.get("ca", 0) or 0
        cogs = proj.get("cogs", 0) or 0
        fw(fin, 138, col, ca - cogs)

    write_proj_row(140, "ca")      # Revenue Products TOTAL (bis)
    write_proj_row(192, "ca")      # Revenue TOTAL
    write_proj_row(194, "cogs")    # COGS TOTAL
    for yk, col in FIN_COL.items():
        proj = proj_by_year.get(yk, {})
        ca = proj.get("ca", 0) or 0
        cogs = proj.get("cogs", 0) or 0
        fw(fin, 195, col, ca - cogs)  # GP TOTAL

    # P&L Statement rows
    write_proj_row(710, "ca")                    # Revenue
    write_proj_row(712, "cogs")                  # COGS
    for yk, col in FIN_COL.items():              # Gross Profit
        proj = proj_by_year.get(yk, {})
        fw(fin, 714, col, (proj.get("ca", 0) or 0) - (proj.get("cogs", 0) or 0))
    write_proj_row(716, "opex_total")            # Operating Expenses
    write_proj_row(718, "ebitda")                 # EBITDA
    write_proj_row(720, "amortissements")         # Amortisations
    write_proj_row(722, "resultat_exploitation")  # EBIT
    write_proj_row(735, "resultat_exploitation")  # Profit before tax (approx)
    write_proj_row(737, "impots", 0)              # Tax regime 1
    write_proj_row(738, "impots", 0)              # Tax regime 2
    write_proj_row(742, "impots")                 # Taxes total
    write_proj_row(744, "resultat_net")           # Net Profit

    # ═══ 16. Cash Flow & Financing (rows 749-799) ═══
    # Row 749 = Cash position at beginning
    write_proj_row(749, "cashflow", 0)

    # Row 733 = Bank charges + Interest total
    # Calculer depuis charges_financieres dans projections
    for yk, col in FIN_COL.items():
        proj = proj_by_year.get(yk, {})
        charges_fin = proj.get("charges_financieres", 0) or 0
        fw(fin, 733, col, charges_fin)

    # Row 729 = Bank charges (petite part des charges financières)
    # Row 730-732 = Interest par source de prêt
    for row in [729, 730, 731, 732]:
        for col in FIN_COL.values():
            cell = fin[f"{col}{row}"]
            if isinstance(cell.value, str) and cell.value.startswith("="):
                fw(fin, row, col, 0)

    # Source of Financing (rows 755-758)
    for row in [755, 756, 757, 758]:
        for col in FIN_COL.values():
            cell = fin[f"{col}{row}"]
            if isinstance(cell.value, str) and cell.value.startswith("="):
                fw(fin, row, col, 0)

    # Interest & Principal in Cash Flow (rows 763-768)
    for row in [763, 764, 765, 768, 769, 770]:
        for col in FIN_COL.values():
            cell = fin[f"{col}{row}"]
            if isinstance(cell.value, str) and cell.value.startswith("="):
                fw(fin, row, col, 0)

    # ═══ 17. Passe finale: écraser TOUTES les formules GETPIVOTDATA restantes ═══
    # Parcourir FinanceData et écraser toute formule contenant GETPIVOTDATA avec 0
    for row in range(1, fin.max_row + 1):
        for col in ['O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X']:
            cell = fin[f"{col}{row}"]
            if isinstance(cell.value, str) and 'GETPIVOTDATA' in cell.value:
                fw(fin, row, col, 0)

    # Aussi dans RevenueData — écraser les GETPIVOTDATA
    for row in range(1, rev.max_row + 1):
        for col in ['O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X']:
            cell = rev[f"{col}{row}"]
            if isinstance(cell.value, str) and 'GETPIVOTDATA' in cell.value:
                fw(rev, row, col, 0)

    # ═══ 18. RevenuePivot BY PRODUCT — remplir le cache PivotTable ═══
    rp = wb["RevenuePivot"]

    # Années dans l'ordre du pivot (N column = year numbers)
    year_labels = ["YEAR-2", "YEAR-1", "CURRENT YEAR", "YEAR2", "YEAR3", "YEAR4", "YEAR5", "YEAR6"]
    year_nums = [years.get(PROJ_YEAR_MAP.get(yl, ""), 0) for yl in year_labels]

    # Colonnes par produit: O=P1, P=P2, Q=P3, R=P4
    prod_cols = ['O', 'P', 'Q', 'R']

    # BY PRODUCT section: rows 15-68
    # VOLUME: rows 15-22 (8 years), row 23 = total
    for yr_idx in range(8):
        row = 15 + yr_idx
        for prod_idx in range(min(len(produits), 4)):
            pa_dict = get_par_annee_dict(produits[prod_idx])
            yl = year_labels[yr_idx]
            yk = PROJ_YEAR_MAP.get(yl, "")
            yr = pa_dict.get(yk, {})
            vol = yr.get("volume", 0) or 0
            fw(rp, row, prod_cols[prod_idx], vol)
        # Produits inactifs = 0
        for prod_idx in range(len(produits), 4):
            fw(rp, row, prod_cols[prod_idx], 0)

    # Total VOLUME (row 23)
    for prod_idx in range(4):
        if prod_idx < len(produits):
            pa_dict = get_par_annee_dict(produits[prod_idx])
            total = sum((pa_dict.get(PROJ_YEAR_MAP.get(yl, ""), {}).get("volume", 0) or 0) for yl in year_labels)
            fw(rp, 23, prod_cols[prod_idx], total)
        else:
            fw(rp, 23, prod_cols[prod_idx], 0)

    # REVENUE: rows 24-31, row 32 = total
    for yr_idx in range(8):
        row = 24 + yr_idx
        for prod_idx in range(min(len(produits), 4)):
            pa_dict = get_par_annee_dict(produits[prod_idx])
            yk = PROJ_YEAR_MAP.get(year_labels[yr_idx], "")
            yr = pa_dict.get(yk, {})
            ca = yr.get("ca", 0) or 0
            if ca == 0:
                vol = yr.get("volume", 0) or 0
                prix = yr.get("prix_r1", produits[prod_idx].get("prix_unitaire", 0)) or 0
                ca = vol * prix
            fw(rp, row, prod_cols[prod_idx], ca)
        for prod_idx in range(len(produits), 4):
            fw(rp, row, prod_cols[prod_idx], 0)

    # Total REVENUE (row 32)
    for prod_idx in range(4):
        if prod_idx < len(produits):
            pa_dict = get_par_annee_dict(produits[prod_idx])
            total = sum((pa_dict.get(PROJ_YEAR_MAP.get(yl, ""), {}).get("ca", 0) or 0) for yl in year_labels)
            fw(rp, 32, prod_cols[prod_idx], total)
        else:
            fw(rp, 32, prod_cols[prod_idx], 0)

    # COST OF GOODS SOLD: rows 33-40, row 41 = total
    for yr_idx in range(8):
        row = 33 + yr_idx
        for prod_idx in range(min(len(produits), 4)):
            pa_dict = get_par_annee_dict(produits[prod_idx])
            yk = PROJ_YEAR_MAP.get(year_labels[yr_idx], "")
            yr = pa_dict.get(yk, {})
            cogs = yr.get("cogs_total", 0) or 0
            if cogs == 0:
                vol = yr.get("volume", 0) or 0
                cogs_u = yr.get("cogs_r1", produits[prod_idx].get("cout_unitaire", 0)) or 0
                cogs = vol * cogs_u
            fw(rp, row, prod_cols[prod_idx], cogs)
        for prod_idx in range(len(produits), 4):
            fw(rp, row, prod_cols[prod_idx], 0)

    # Total COGS (row 41)
    for prod_idx in range(4):
        if prod_idx < len(produits):
            pa_dict = get_par_annee_dict(produits[prod_idx])
            total = sum((pa_dict.get(PROJ_YEAR_MAP.get(yl, ""), {}).get("cogs_total", 0) or 0) for yl in year_labels)
            fw(rp, 41, prod_cols[prod_idx], total)
        else:
            fw(rp, 41, prod_cols[prod_idx], 0)

    # GROSS PROFIT: rows 42-49, row 50 = total (GETPIVOTDATA formulas → replace with values)
    for yr_idx in range(8):
        row = 42 + yr_idx
        for prod_idx in range(min(len(produits), 4)):
            pa_dict = get_par_annee_dict(produits[prod_idx])
            yk = PROJ_YEAR_MAP.get(year_labels[yr_idx], "")
            yr = pa_dict.get(yk, {})
            ca = yr.get("ca", 0) or 0
            cogs = yr.get("cogs_total", 0) or 0
            fw(rp, row, prod_cols[prod_idx], ca - cogs)
        for prod_idx in range(len(produits), 4):
            fw(rp, row, prod_cols[prod_idx], 0)

    # Total GP (row 50)
    for prod_idx in range(4):
        if prod_idx < len(produits):
            pa_dict = get_par_annee_dict(produits[prod_idx])
            total_ca = sum((pa_dict.get(PROJ_YEAR_MAP.get(yl, ""), {}).get("ca", 0) or 0) for yl in year_labels)
            total_cogs = sum((pa_dict.get(PROJ_YEAR_MAP.get(yl, ""), {}).get("cogs_total", 0) or 0) for yl in year_labels)
            fw(rp, 50, prod_cols[prod_idx], total_ca - total_cogs)
        else:
            fw(rp, 50, prod_cols[prod_idx], 0)

    # GROSS PROFIT %: rows 51-59 — replace GETPIVOTDATA formulas with calculated %
    for yr_idx in range(8):
        row = 51 + yr_idx
        for prod_idx in range(min(len(produits), 4)):
            pa_dict = get_par_annee_dict(produits[prod_idx])
            yk = PROJ_YEAR_MAP.get(year_labels[yr_idx], "")
            yr = pa_dict.get(yk, {})
            ca = yr.get("ca", 0) or 0
            cogs = yr.get("cogs_total", 0) or 0
            gp_pct = (ca - cogs) / ca if ca > 0 else 0
            fw(rp, row, prod_cols[prod_idx], round(gp_pct, 4))
        for prod_idx in range(len(produits), 4):
            fw(rp, row, prod_cols[prod_idx], 0)
    # Total GP % (row 59)
    for prod_idx in range(4):
        if prod_idx < len(produits):
            pa_dict = get_par_annee_dict(produits[prod_idx])
            total_ca = sum((pa_dict.get(PROJ_YEAR_MAP.get(yl, ""), {}).get("ca", 0) or 0) for yl in year_labels)
            total_cogs = sum((pa_dict.get(PROJ_YEAR_MAP.get(yl, ""), {}).get("cogs_total", 0) or 0) for yl in year_labels)
            fw(rp, 59, prod_cols[prod_idx], round((total_ca - total_cogs) / total_ca, 4) if total_ca > 0 else 0)
        else:
            fw(rp, 59, prod_cols[prod_idx], 0)

    # BY QUARTER section (left side, cols D-J): write from product par_annee quarterly data
    # D=ITEM label, E=YEAR, F=Q1, G=Q2, H=Q3, I=Q4, J=TOTAL
    # VOLUME: rows 16-23 (row 15 = header, row 24 = total)
    for yr_idx in range(8):
        row = 16 + yr_idx
        # Aggregate across all products
        total_q1, total_q2, total_q3, total_q4 = 0, 0, 0, 0
        for p in produits:
            pa_dict = get_par_annee_dict(p)
            yk = PROJ_YEAR_MAP.get(year_labels[yr_idx], "")
            yr = pa_dict.get(yk, {})
            vol = yr.get("volume", 0) or 0
            q1 = yr.get("volume_q1", 0) or round(vol * 0.22)
            q2 = yr.get("volume_q2", 0) or round(vol * 0.25)
            q3 = yr.get("volume_q3", 0) or round(vol * 0.27)
            q4 = yr.get("volume_q4", 0) or (vol - q1 - q2 - q3)
            total_q1 += q1; total_q2 += q2; total_q3 += q3; total_q4 += q4
        fw(rp, row, "F", total_q1)
        fw(rp, row, "G", total_q2)
        fw(rp, row, "H", total_q3)
        fw(rp, row, "I", total_q4)
        fw(rp, row, "J", total_q1 + total_q2 + total_q3 + total_q4)

    # Total VOLUME row 24
    grand_q1 = sum(rp[f"F{16+i}"].value or 0 for i in range(8))
    grand_q2 = sum(rp[f"G{16+i}"].value or 0 for i in range(8))
    grand_q3 = sum(rp[f"H{16+i}"].value or 0 for i in range(8))
    grand_q4 = sum(rp[f"I{16+i}"].value or 0 for i in range(8))
    fw(rp, 24, "F", grand_q1); fw(rp, 24, "G", grand_q2)
    fw(rp, 24, "H", grand_q3); fw(rp, 24, "I", grand_q4)
    fw(rp, 24, "J", grand_q1 + grand_q2 + grand_q3 + grand_q4)

    # ═══ FINAL ═══
    wb.calculation = openpyxl.workbook.properties.CalcProperties(fullCalcOnLoad=True)

    n_prod = len(produits)
    n_svc = len(services)
    n_staff = len(staff_list)
    n_capex = sum(len(v) for v in categorized.values() if isinstance(v, list))
    print(f"[ovo_filler] Done: {n_prod} produits, {n_svc} services, "
          f"{n_staff} staff, {n_capex} capex")
    return wb


def _fill_default_scenarios(inp, avg_growth):
    """Scénarios par défaut en multiplicateurs (comme le template attend)."""
    rev_w = round(1 - avg_growth * 0.3, 3)
    rev_b = round(1 + avg_growth * 0.3, 3)
    cogs_w, cogs_b = 1.05, 0.97
    opex_w, opex_b = 1.10, 0.90

    for row in [130, 132]:
        fw(inp, row, "H", rev_w); fw(inp, row, "I", 1.0); fw(inp, row, "J", rev_b)
    for row in [131, 133]:
        fw(inp, row, "H", cogs_w); fw(inp, row, "I", 1.0); fw(inp, row, "J", cogs_b)
    for row in [134, 135, 136, 137, 138, 139, 140, 141, 142]:
        fw(inp, row, "H", opex_w); fw(inp, row, "I", 1.0); fw(inp, row, "J", opex_b)
